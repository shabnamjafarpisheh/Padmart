import streamlit as st
from datetime import date, datetime, timedelta
import uuid, hashlib, io, time
import plotly.graph_objects as pgo   # renamed alias — avoids collision with nav helper
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ─────────────────────────────────────────────
#  PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Padmart — Acting School",
    page_icon="🎭",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ─────────────────────────────────────────────
#  CONSTANTS
# ─────────────────────────────────────────────
HALLS = {
    "grand":   {"name_en": "Grand Studio", "name_fa": "پلاتو بزرگ",
                "price": 70000, "icon": "🎪",
                "desc_en": " · Pro lighting · ",
                "desc_fa": "  · نور حرفه‌ای ·  ",
                "color": "#6d1f2a", "light": "rgba(109,31,42,0.15)"},
    "studio2": {"name_en": "Studio 2",     "name_fa": "پلاتو ۲",
                "price": 70000,   "icon": "🎭",
                "desc_en": "  · Mirrors · Sound system",
                "desc_fa": "  · آینه · سیستم صوتی",
                "color": "#7c6fcd", "light": "rgba(124,111,205,0.15)"},
}
SLOT_STARTS = list(range(9, 21))   # 09:00 → 20:00 (last slot ends 21:00)
OWNER_PASS_DEFAULT = "admin123"
WINE  = "#6d1f2a"
GOLD  = "#b88a4a"
CREAM = "#f5eadf"

# ─────────────────────────────────────────────
#  SESSION STATE
# ─────────────────────────────────────────────
def _init():
    defs = {
        "lang":         "en",
        "page":         "gateway",
        "owner_auth":   False,
        "owner_pass":   OWNER_PASS_DEFAULT,
        "reservations": [],
        "users":        {},      # {email: {name, phone, pass_hash}}
        "current_user": None,
        "smtp":         {"host":"","port":587,"user":"","pass":""},
    }
    for k, v in defs.items():
        if k not in st.session_state:
            st.session_state[k] = v
_init()

# ─────────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────────
def hpw(pw):   return hashlib.sha256(pw.encode()).hexdigest()
def fmt(n):    return f"{n:,} تومان"
def gen_id():  return "RES-" + str(uuid.uuid4())[:8].upper()
def slot_label(s, d): return f"{s:02d}:00–{s+d:02d}:00"
def is_fa():   return st.session_state.lang == "fa"

def cancellation_fee(r):
    """
    Returns (fee_amount, is_free).
    If cancelled 48h+ before reservation start → free (fee=0).
    If cancelled < 48h before → 100% charge.
    """
    res_start = datetime.strptime(r["date"] + f" {r['start_hour']:02d}:00", "%Y-%m-%d %H:%M")
    hours_left = (res_start - datetime.now()).total_seconds() / 3600
    if hours_left >= 48:
        return 0, True
    else:
        return r["price"], False

def build_excel(reservations, users):
    """Build an Excel workbook with reservation report for owner."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reservations"

    wine_fill   = PatternFill("solid", fgColor="6D1F2A")
    gold_fill   = PatternFill("solid", fgColor="B88A4A")
    cream_fill  = PatternFill("solid", fgColor="FFF8F0")
    cancel_fill = PatternFill("solid", fgColor="F2E2E0")
    hdr_font    = Font(bold=True, color="FFFFFF", size=11)
    bold_wine   = Font(bold=True, color="6D1F2A", size=10)
    normal_font = Font(size=10)
    center      = Alignment(horizontal="center", vertical="center")
    thin        = Border(
        left=Side(style="thin", color="E7D8C6"),
        right=Side(style="thin", color="E7D8C6"),
        top=Side(style="thin", color="E7D8C6"),
        bottom=Side(style="thin", color="E7D8C6"),
    )

    headers = [
        "Reservation ID", "Full Name", "Phone", "Email",
        "Hall", "Date", "Slot", "Duration (h)", "Total (تومان)", "Status", "Booked At"
    ]
    col_w = [14, 22, 16, 28, 16, 12, 16, 14, 18, 12, 18]

    ws.merge_cells("A1:K1")
    t = ws["A1"]
    t.value = "Padmart — Reservation Report"
    t.font  = Font(bold=True, color="FFFFFF", size=14)
    t.fill  = wine_fill
    t.alignment = center
    ws.row_dimensions[1].height = 28

    for ci, (h, w) in enumerate(zip(headers, col_w), start=1):
        c = ws.cell(row=2, column=ci, value=h)
        c.font      = hdr_font
        c.fill      = gold_fill
        c.alignment = center
        c.border    = thin
        ws.column_dimensions[c.column_letter].width = w
    ws.row_dimensions[2].height = 20

    for ri, r in enumerate(sorted(reservations, key=lambda x: x["date"], reverse=True), start=3):
        phone = users.get(r["email"], {}).get("phone", "—")
        row_fill = cancel_fill if r["status"] == "cancelled" else cream_fill
        vals = [
            r["id"], r["name"], phone, r["email"],
            r.get("hall_en", r["hall"]), r["date"], r["slot_label"],
            r["duration"], r["price"], r["status"].capitalize(), r.get("booked_at","")
        ]
        for ci, v in enumerate(vals, start=1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.fill      = row_fill
            c.font      = bold_wine if ci == 9 else normal_font
            c.alignment = center
            c.border    = thin
        ws.row_dimensions[ri].height = 18

    active = [r for r in reservations if r["status"] == "active"]
    tot_row = len(reservations) + 3
    ws.merge_cells(f"A{tot_row}:H{tot_row}")
    s = ws.cell(row=tot_row, column=1, value="TOTAL ACTIVE REVENUE")
    s.font = Font(bold=True, color="6D1F2A", size=11)
    s.fill = PatternFill("solid", fgColor="F5EADF")
    s.alignment = Alignment(horizontal="right", vertical="center")
    t2 = ws.cell(row=tot_row, column=9, value=sum(r["price"] for r in active))
    t2.font = Font(bold=True, color="6D1F2A", size=11)
    t2.fill = PatternFill("solid", fgColor="F5EADF")
    t2.alignment = center
    ws.row_dimensions[tot_row].height = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

def booked_hours(hall, date_str):
    taken = set()
    for r in st.session_state.reservations:
        if r["hall"]==hall and r["date"]==date_str and r["status"]=="active":
            for h in range(r["start_hour"], r["start_hour"]+r["duration"]):
                taken.add(h)
    return taken

def can_book(hall, date_str, start, dur):
    taken = booked_hours(hall, date_str)
    for h in range(start, start+dur):
        if h in taken or h > 20: return False
    return True

def cancel_res(res_id):
    for r in st.session_state.reservations:
        if r["id"] == res_id:
            r["status"] = "cancelled"
            break

def nav(page):
    st.session_state.page = page
    st.rerun()

# ─────────────────────────────────────────────
#  GLOBAL CSS
# ─────────────────────────────────────────────
FA   = is_fa()
FONT = "'Vazirmatn','Tahoma',sans-serif" if FA else "'DM Sans',sans-serif"
DIR  = "rtl" if FA else "ltr"

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=Playfair+Display:wght@600;700&family=Vazirmatn:wght@400;500;600;700&display=swap');
:root{{--wine:#6d1f2a;--gold:#b88a4a;--muted:#7f6b62;--line:#e7d8c6;--paper:#fffaf3;--ink:#241817}}
html,body,[data-testid="stAppViewContainer"]{{
  background:linear-gradient(135deg,#f6ecdf 0%,#fbf7f0 50%,#f1e2d2 100%)!important;
  font-family:{FONT}!important;color:var(--ink);direction:{DIR};}}
[data-testid="stHeader"]{{background:rgba(255,249,240,.96)!important;border-bottom:1px solid var(--line)}}
[data-testid="stSidebar"]{{display:none}}
#MainMenu,footer,[data-testid="stToolbar"]{{display:none!important}}
section[data-testid="stMain"]{{padding-top:0!important}}
.display{{font-family:'Playfair Display',serif;font-weight:700}}

.hero{{background:linear-gradient(90deg,rgba(36,24,23,.86),rgba(36,24,23,.28) 60%,rgba(36,24,23,.06)),
  url('https://images.unsplash.com/photo-1503095396549-807759245b35?w=1400&q=80&fit=crop') center/cover;
  border-radius:24px;padding:48px 44px;color:#fff;margin-bottom:22px;
  box-shadow:0 16px 48px rgba(72,16,25,.18);}}
.hero h1{{font-family:'Playfair Display',serif;font-size:34px;font-weight:700;line-height:1.1;margin:0 0 10px}}
.hero p{{font-size:14px;color:rgba(255,255,255,.84);line-height:1.7;margin:0;max-width:500px}}
.hero-badge{{display:inline-flex;align-items:center;gap:6px;padding:6px 14px;border-radius:99px;
  background:var(--gold);font-size:11px;font-weight:600;letter-spacing:.18em;
  text-transform:uppercase;color:#fff;margin-bottom:14px}}
.owner-hero{{background:linear-gradient(90deg,rgba(15,5,8,.92),rgba(15,5,8,.42) 60%,rgba(15,5,8,.08)),
  url('https://images.unsplash.com/photo-1503095396549-807759245b35?w=1400&q=80&fit=crop') center/cover;
  border-radius:24px;padding:36px 44px;color:#fff;margin-bottom:22px;
  box-shadow:0 16px 48px rgba(72,16,25,.22);}}
.owner-hero h1{{font-family:'Playfair Display',serif;font-size:28px;font-weight:700;margin:0 0 6px}}
.owner-hero p{{font-size:13px;color:rgba(255,255,255,.78);margin:0}}
.owner-badge{{display:inline-flex;align-items:center;gap:4px;
  background:linear-gradient(135deg,var(--gold),#8a6030);color:#fff;
  font-size:11px;font-weight:700;letter-spacing:.12em;text-transform:uppercase;
  padding:3px 10px;border-radius:99px;margin-bottom:10px}}

.panel{{background:var(--paper);border:1px solid var(--line);border-radius:20px;
  padding:22px;box-shadow:0 6px 24px rgba(72,16,25,.05);margin-bottom:14px}}
.panel-dark{{background:#2f1d1a;border-radius:20px;padding:22px;color:#fff;margin-bottom:14px}}
.eyebrow{{font-size:10px;font-weight:600;letter-spacing:.22em;text-transform:uppercase;color:#9b7250;margin-bottom:4px}}
.eyebrow-light{{color:#d8b47b}}
.gcard{{border-radius:24px;padding:34px 22px;text-align:center;
  border:1.5px solid var(--line);background:var(--paper);margin-bottom:10px}}
.gcard-owner{{background:#1e0f0f;border:2px solid var(--gold);border-radius:24px;
  padding:34px 22px;text-align:center;margin-bottom:10px}}
.hall-card{{border:1px solid var(--line);border-radius:16px;padding:16px;
  background:#fffdf9;margin-bottom:8px}}
.hall-card.sel{{border-color:var(--wine);background:#fff7f7;
  box-shadow:0 6px 20px rgba(109,31,42,.1)}}
.stat-card{{background:var(--paper);border:1px solid var(--line);border-radius:18px;
  padding:18px;text-align:center;box-shadow:0 4px 14px rgba(72,16,25,.04)}}
.stat-num{{font-family:'Playfair Display',serif;font-size:30px;color:var(--wine);font-weight:700}}
.stat-lbl{{font-size:10px;font-weight:600;letter-spacing:.14em;text-transform:uppercase;
  color:#9b7250;margin-top:5px}}
.res-card{{background:var(--paper);border:1px solid var(--line);border-radius:18px;
  padding:16px;margin-bottom:10px;box-shadow:0 4px 12px rgba(72,16,25,.04)}}
.res-card.cancelled{{opacity:.6}}
.res-id{{font-size:10px;font-weight:600;letter-spacing:.15em;text-transform:uppercase;
  color:#9b7250;margin-bottom:3px}}
.res-hall{{font-family:'Playfair Display',serif;font-size:18px;color:#2b1c19;margin-bottom:4px}}
.res-price{{font-family:'Playfair Display',serif;font-size:20px;color:var(--wine)}}
.badge-active{{display:inline-flex;padding:3px 10px;border-radius:99px;
  font-size:11px;font-weight:600;background:#e7f3ec;color:#2d7a58}}
.badge-cancelled{{display:inline-flex;padding:3px 10px;border-radius:99px;
  font-size:11px;font-weight:600;background:#f2e2e0;color:#8b2d36}}
.badge-owner{{display:inline-flex;padding:3px 10px;border-radius:99px;
  font-size:11px;font-weight:600;background:rgba(184,138,74,.18);color:#7a5a22}}
.summary-strip{{background:#f5eadf;border-radius:12px;padding:12px 16px;
  display:flex;justify-content:space-between;align-items:center;margin:12px 0}}
.msg-ok{{background:#e8f3ec;color:#2d6b4f;border-radius:11px;
  padding:10px 14px;font-size:14px;font-weight:500;margin:8px 0}}
.msg-err{{background:#f8e5e3;color:#8b2d36;border-radius:11px;
  padding:10px 14px;font-size:14px;font-weight:500;margin:8px 0}}
.msg-info{{background:#e8f0fb;color:#1a4a8a;border-radius:11px;
  padding:10px 14px;font-size:14px;font-weight:500;margin:8px 0}}

.stTextInput>div>div>input,.stNumberInput>div>div>input,.stDateInput>div>div>input{{
  border:1px solid var(--line)!important;background:#fffdf9!important;
  border-radius:12px!important;padding:12px 14px!important;
  font-family:{FONT}!important;font-size:14px!important;color:var(--ink)!important}}
.stTextInput>div>div>input:focus{{
  border-color:var(--wine)!important;box-shadow:0 0 0 3px rgba(109,31,42,.1)!important}}
label[data-testid="stWidgetLabel"]>div>p{{
  font-size:13px!important;font-weight:600!important;color:var(--ink)!important}}
.stSelectbox>div>div>div{{border:1px solid var(--line)!important;background:#fffdf9!important;
  border-radius:12px!important;font-family:{FONT}!important}}
.stButton>button{{
  background:var(--wine)!important;color:#fff!important;border:none!important;
  border-radius:13px!important;padding:11px 20px!important;
  font-family:{FONT}!important;font-size:14px!important;font-weight:600!important;
  box-shadow:0 4px 14px rgba(109,31,42,.2)!important;transition:.2s!important;width:100%!important}}
.stButton>button:hover{{background:#571821!important;transform:translateY(-1px)!important}}
.btn-gold>button{{background:linear-gradient(135deg,var(--gold),#8a6030)!important;
  color:#1e0f0f!important;box-shadow:0 4px 14px rgba(184,138,74,.25)!important}}
.btn-cancel>button{{background:#f2e2e0!important;color:#7b2932!important;
  box-shadow:none!important;font-size:12px!important;padding:7px 14px!important}}
.btn-ghost>button{{background:none!important;color:var(--muted)!important;
  border:1px solid var(--line)!important;box-shadow:none!important;font-size:13px!important}}
.stTabs [data-baseweb="tab-list"]{{gap:4px;background:transparent;border-bottom:2px solid var(--line)}}
.stTabs [data-baseweb="tab"]{{background:transparent!important;border:none!important;
  color:var(--muted)!important;font-family:{FONT}!important;font-weight:600!important;
  font-size:13px!important;padding:9px 16px!important;border-radius:0!important}}
.stTabs [aria-selected="true"]{{color:var(--wine)!important;border-bottom:2px solid var(--wine)!important}}
.stTabs [data-baseweb="tab-panel"]{{padding-top:18px!important}}
hr{{border:none;border-top:1px solid var(--line);margin:14px 0}}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────
#  LANGUAGE TOGGLE
# ─────────────────────────────────────────────
_, lc = st.columns([7, 1])
with lc:
    lx = st.selectbox("🌐", ["English","فارسی"],
                      index=0 if st.session_state.lang=="en" else 1,
                      label_visibility="collapsed", key="lang_sel")
    nl = "en" if lx=="English" else "fa"
    if nl != st.session_state.lang:
        st.session_state.lang = nl
        st.rerun()

# ─────────────────────────────────────────────
#  CALENDAR CHART
#  X-axis = days (dates), Y-axis = hours (time)
#  Select hall first; each cell shows booked / free
# ─────────────────────────────────────────────
def hall_calendar(
    hall_key: str,
    days_ahead: int = 14,
    highlight_email: str = None,
    allow_booking: bool = False,
    booking_email: str = None,
    booking_name: str = None,
    is_owner: bool = False,
    cal_prefix: str = "cal",
):
    """
    Heatmap calendar for ONE hall.
    X = days (shown at TOP), Y = hours (09:00–20:00, reversed so morning is at top).
    Cells: FREE (cream) | BOOKED-other (hall colour) | MY-booking (gold)

    When allow_booking=True, a 'Book from Calendar' section appears below
    the chart so users can click a date + start hour to pre-fill a booking.
    """
    FA       = is_fa()
    hall     = HALLS[hall_key]
    hall_col = hall["color"]
    hall_nm  = hall["name_fa"] if FA else hall["name_en"]

    today_     = date.today()
    date_range = [str(today_ + timedelta(days=i)) for i in range(days_ahead)]
    hours      = list(range(9, 21))   # 09 → 20

    # ── collect booked hours ──
    booked = {}   # (date_str, hour_int) -> dict
    for r in st.session_state.reservations:
        if r["hall"] == hall_key and r["status"] == "active":
            for h in range(r["start_hour"], r["start_hour"] + r["duration"]):
                booked[(r["date"], h)] = {
                    "name":  r["name"],
                    "email": r["email"],
                    "slot":  r["slot_label"],
                    "price": fmt(r["price"]),
                    "mine":  r["email"] == highlight_email if highlight_email else False,
                }

    # ── build z + hover matrices ──
    z_vals  = []
    hover   = []
    y_lbl   = [f"{h:02d}:00" for h in hours]

    for h in hours:
        z_row, h_row = [], []
        for d in date_range:
            info = booked.get((d, h))
            if info is None:
                z_row.append(0)
                h_row.append(
                    "<b>" + ("Free — click to book" if not FA else "آزاد — کلیک کنید") + "</b><br>"
                    + d + " " + f"{h:02d}:00"
                )
            elif info["mine"]:
                z_row.append(2)
                h_row.append(
                    "<b>" + ("Your Booking" if not FA else "رزرو شما") + "</b><br>"
                    + d + " " + info["slot"] + "<br>"
                    + ("Guest" if not FA else "مهمان") + ": " + info["name"] + "<br>"
                    + info["price"]
                )
            else:
                z_row.append(1)
                h_row.append(
                    "<b>" + ("Booked" if not FA else "رزرو شده") + "</b><br>"
                    + d + " " + info["slot"] + "<br>"
                    + ("Guest" if not FA else "مهمان") + ": " + info["name"] + "<br>"
                    + info["price"]
                )
        z_vals.append(z_row)
        hover.append(h_row)

    x_lbl = []
    for d in date_range:
        dt = datetime.strptime(d, "%Y-%m-%d")
        x_lbl.append(dt.strftime("%a %d"))

    colorscale = [
        [0.000, "#f0e8db"],
        [0.499, "#f0e8db"],
        [0.500, hall_col],
        [0.850, hall_col],
        [0.851, GOLD],
        [1.000, GOLD],
    ]

    fig = pgo.Figure(pgo.Heatmap(
        z=z_vals, x=x_lbl, y=y_lbl,
        colorscale=colorscale, zmin=0, zmax=2,
        showscale=False,
        hovertemplate="%{customdata}<extra></extra>",
        customdata=hover,
        xgap=4, ygap=4,
    ))

    # text annotations inside booked cells
    annotations = []
    for yi, h in enumerate(hours):
        for xi, d in enumerate(date_range):
            info = booked.get((d, h))
            if info:
                short = (info["name"].split()[0] if info["name"] else "")[:9]
                annotations.append(dict(
                    x=x_lbl[xi], y=f"{h:02d}:00",
                    text="<b>" + short + "</b>",
                    font=dict(size=9, color="white"),
                    showarrow=False, xref="x", yref="y",
                ))

    fig.update_layout(
        title=dict(
            text=hall["icon"] + " " + hall_nm + " — " + ("Availability Calendar" if not FA else "جدول زمانبندی"),
            font=dict(family="Playfair Display, serif", size=17, color=WINE),
            x=0.5,
        ),
        xaxis=dict(
            side="top",
            tickfont=dict(size=12, color="#241817", family="DM Sans"),
            tickangle=0,
            showgrid=False, fixedrange=True,
            title=dict(text=""),
        ),
        yaxis=dict(
            tickfont=dict(size=12, color="#241817", family="DM Sans"),
            showgrid=False, autorange="reversed", fixedrange=True,
            title=dict(text=""),
        ),
        plot_bgcolor="#fffaf3", paper_bgcolor="#fffaf3",
        margin=dict(l=10, r=10, t=80, b=10),
        height=max(380, 50 + len(hours) * 32),
        annotations=annotations,
        hoverlabel=dict(
            bgcolor="#fffaf3", bordercolor=WINE,
            font=dict(family="DM Sans", size=12, color="#241817"),
        ),
    )

    # legend
    my_legend = (
        '<div style="display:inline-flex;align-items:center;gap:6px;margin-left:16px">'
        '<div style="width:14px;height:14px;border-radius:3px;background:' + GOLD + '"></div>'
        '<span style="color:var(--muted)">' + ("Your reservation" if not FA else "رزرو شما") + '</span>'
        '</div>'
    ) if highlight_email else ""

    st.markdown(
        '<div style="display:flex;gap:6px;align-items:center;margin-bottom:8px;font-size:12px;flex-wrap:wrap">'
        '<div style="display:inline-flex;align-items:center;gap:6px">'
        '<div style="width:14px;height:14px;border-radius:3px;background:#f0e8db;border:1px solid #e7d8c6"></div>'
        '<span style="color:var(--muted)">' + ("Free" if not FA else "آزاد") + '</span>'
        '</div>'
        '<div style="display:inline-flex;align-items:center;gap:6px;margin-left:16px">'
        '<div style="width:14px;height:14px;border-radius:3px;background:' + hall_col + '"></div>'
        '<span style="color:var(--muted)">' + ("Booked" if not FA else "رزرو شده") + '</span>'
        '</div>'
        + my_legend +
        '</div>',
        unsafe_allow_html=True
    )

    st.plotly_chart(fig, use_container_width=True)

    # ── BOOK FROM CALENDAR ──
    if not allow_booking:
        return

    st.markdown(
        '<div style="background:#f5eadf;border-radius:16px;padding:18px 20px;margin-top:4px">'
        '<div style="font-size:10px;font-weight:600;letter-spacing:.22em;text-transform:uppercase;color:#9b7250;margin-bottom:8px">'
        + ("Book from Calendar" if not FA else "رزرو از تقویم") +
        '</div>'
        '<div style="font-size:13px;color:#5e4d46;margin-bottom:14px">'
        + ("Pick a date and start hour from the free (cream) cells above, set your duration, then confirm."
           if not FA else
           "یک تاریخ و ساعت شروع از زمان آزاد (کرم رنگ) پایین انتخاب کنید، مدت زمان را تنظیم کنید و تأیید کنید.")
        + '</div>'
        '</div>',
        unsafe_allow_html=True
    )

    with st.form(f"cal_book_{cal_prefix}_{hall_key}"):
        if is_owner:
            cc1, cc2 = st.columns(2)
            with cc1:
                gname  = st.text_input("Guest Name"            if not FA else "نام مهمان",             key=f"cb_{cal_prefix}_gn")
            with cc2:
                gemail = st.text_input("Guest Email (optional)" if not FA else "ایمیل مهمان (اختیاری)", key=f"cb_{cal_prefix}_ge")
            gphone = st.text_input("Guest Phone" if not FA else "تلفن مهمان", key=f"cb_{cal_prefix}_gph", placeholder="+98 912 345 6789")
        else:
            gname, gemail = booking_name, booking_email
            gphone = st.session_state.users.get(booking_email, {}).get("phone", "")

        # date picker — only future dates
        today_v = date.today()
        cb_date = st.date_input(
            "Date" if not FA else "تاریخ",
            min_value=today_v,
            max_value=today_v + timedelta(days=days_ahead - 1),
            value=today_v + timedelta(days=1),
            key=f"cb_{cal_prefix}_date"
        )
        cb_dur = st.slider(
            "Duration (hours)" if not FA else "مدت زمان (ساعت)",
            1, 6, 1,
            key=f"cb_{cal_prefix}_dur"
        )

        cb_date_str = str(cb_date)
        avail_starts = [
            s for s in SLOT_STARTS
            if s + cb_dur <= 21 and can_book(hall_key, cb_date_str, s, cb_dur)
        ]

        if avail_starts:
            slot_map = {slot_label(s, cb_dur): s for s in avail_starts}
            cb_slot  = st.selectbox(
                "Start Time" if not FA else "ساعت شروع",
                list(slot_map.keys()),
                key=f"cb_{cal_prefix}_slot"
            )
            cb_start = slot_map[cb_slot]
        else:
            st.warning(
                "No available slots for this date / duration." if not FA
                else "هیچ بازه‌ای برای این تاریخ / زمان موجود نیست."
            )
            cb_start = None

        total = HALLS[hall_key]["price"] * cb_dur
        st.markdown(
            '<div style="background:rgba(109,31,42,.07);border-radius:10px;padding:10px 14px;'
            'display:flex;justify-content:space-between;align-items:center;margin-top:6px">'
            '<span style="font-size:13px;color:#5e4d46">'
            + hall["icon"] + " " + hall_nm + " · " + (slot_label(cb_start, cb_dur) if cb_start else "—") +
            '</span>'
            '<span style="font-family:\'Playfair Display\',serif;font-size:18px;color:' + WINE + '">'
            + (fmt(total) if cb_start else "—") +
            '</span>'
            '</div>',
            unsafe_allow_html=True
        )

        confirmed = st.form_submit_button(
            "🎫 Confirm Reservation" if not FA else "🎫 تأیید رزرو",
            use_container_width=True
        )

    if confirmed:
        actual_name  = (gname.strip()  or "Owner Booking")      if is_owner else booking_name
        actual_email = (gemail.strip().lower() or "owner@padmart.internal") if is_owner else booking_email
        actual_phone = gphone.strip() if is_owner else st.session_state.users.get(booking_email, {}).get("phone","")

        if cb_start is None:
            st.markdown('<div class="msg-err">⚠ ' + ("No slot selected." if not FA else "زمانی انتخاب نشده.") + "</div>", unsafe_allow_html=True)
        elif not can_book(hall_key, cb_date_str, cb_start, cb_dur):
            st.markdown('<div class="msg-err">⚠ ' + ("Slot just taken." if not FA else "زمانی همین لحظه رزرو شد.") + "</div>", unsafe_allow_html=True)
        elif (not is_owner and
              len([r for r in st.session_state.reservations
                   if r["email"] == actual_email and r["status"] == "active"]) >= 5):
            st.markdown('<div class="msg-err">⚠ ' + ("Max 5 active bookings." if not FA else "حداکثر ۵ رزرو فعال.") + "</div>", unsafe_allow_html=True)
        else:
            sl  = slot_label(cb_start, cb_dur)
            new = {
                "id": gen_id(), "name": actual_name, "email": actual_email, "phone": actual_phone,
                "hall": hall_key, "hall_en": hall["name_en"], "hall_fa": hall["name_fa"],
                "hall_color": hall["color"],
                "date": cb_date_str, "start_hour": cb_start, "duration": cb_dur,
                "slot_label": sl, "price": total, "status": "active",
                "booked_by": "owner" if is_owner else "user",
                "booked_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            }
            st.session_state.reservations.append(new)
            st.markdown(
                '<div class="msg-ok">✓ ' +
                ("Confirmed! " + hall_nm + " · " + sl if not FA else "تأیید شد! " + hall_nm + " · " + sl) +
                '</div>',
                unsafe_allow_html=True
            )
            st.balloons()
            time.sleep(3)
            st.rerun()


# ─────────────────────────────────────────────
#  BOOKING FORM (shared: users + owner)
# ─────────────────────────────────────────────
def booking_form(prefix, by_email, by_name, is_owner=False):
    FA = is_fa()
    with st.form(f"book_{prefix}"):
        if is_owner:
            st.markdown(f'<div class="msg-info">👑 {"Booking as owner." if not FA else "رزرو به عنوان مدیر."}</div>',
                        unsafe_allow_html=True)
            gname  = st.text_input("Guest Name"            if not FA else "نام مهمان",  key=f"{prefix}_gn")
            gphone = st.text_input("Guest Phone"           if not FA else "تلفن مهمان", key=f"{prefix}_gph", placeholder="+98 912 345 6789")
            gemail = st.text_input("Guest Email (optional)" if not FA else "ایمیل مهمان (اختیاری)", key=f"{prefix}_ge")
        else:
            gname  = by_name
            gemail = by_email

        bdate = st.date_input("Date" if not FA else "تاریخ",
                              min_value=date.today(),
                              max_value=date.today()+timedelta(days=60),
                              value=date.today()+timedelta(days=1),
                              key=f"{prefix}_d")

        hall_options = {(HALLS[k]["name_fa"] if FA else HALLS[k]["name_en"]): k
                        for k in HALLS}
        h_label = st.selectbox("Choose a Hall" if not FA else "انتخاب پلاتو",
                               list(hall_options.keys()), key=f"{prefix}_h")
        hall_key = hall_options[h_label]
        hall     = HALLS[hall_key]

        dur = st.slider("Duration (hours)" if not FA else "مدت زمان (ساعت)",
                        1, 6, 1, key=f"{prefix}_dur")

        date_str = str(bdate)
        avail = [s for s in SLOT_STARTS if s+dur<=21 and can_book(hall_key, date_str, s, dur)]

        if avail:
            slot_map = {slot_label(s, dur): s for s in avail}
            chosen   = st.selectbox("Start Time Slot" if not FA else "زمان شروع",
                                    list(slot_map.keys()), key=f"{prefix}_sl")
            start_hr = slot_map[chosen]
        else:
            st.warning("No available slots for this hall / date / duration." if not FA
                       else "هیچ وقت آزادی موجود نیست.")
            start_hr = None

        total    = hall["price"] * dur
        hall_nm  = hall["name_fa"] if FA else hall["name_en"]
        slot_d   = slot_label(start_hr, dur) if start_hr is not None else "—"

        st.markdown(f"""
        <div class="summary-strip">
          <div>
            <div style="font-size:10px;font-weight:600;letter-spacing:.14em;
                        text-transform:uppercase;color:#9b7250">
              {"Selected Hall" if not FA else "پلاتو انتخابی"}
            </div>
            <div style="font-size:13px;color:#5e4d46;margin-top:2px">
              {hall["icon"]} {hall_nm} · {slot_d}
            </div>
          </div>
          <div style="text-align:{'left' if FA else 'right'}">
            <div style="font-size:10px;font-weight:600;letter-spacing:.14em;
                        text-transform:uppercase;color:#9b7250">
              {"Total Price" if not FA else "قیمت کل"}
            </div>
            <div style="font-family:'Playfair Display',serif;font-size:20px;
                        color:var(--wine);margin-top:2px">
              {fmt(total) if start_hr is not None else "—"}
            </div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        submitted = st.form_submit_button(
            "🎫 Confirm Reservation & Pay" if not FA else "🎫 تأیید رزرو و پرداخت در محل",
            use_container_width=True
        )

    if submitted:
        actual_name  = (gname.strip()  or "Owner Booking") if is_owner else gname
        actual_email = (gemail.strip().lower() or "owner@padmart.internal") if is_owner else gemail
        actual_phone = gphone.strip() if is_owner else st.session_state.users.get(by_email, {}).get("phone","")

        if start_hr is None:
            st.markdown('<div class="msg-err">⚠ ' +
                        ("No slot selected." if not FA else "زمانی انتخاب نشده.") +
                        "</div>", unsafe_allow_html=True)
        elif not can_book(hall_key, date_str, start_hr, dur):
            st.markdown('<div class="msg-err">⚠ ' +
                        ("Slot just taken — pick another." if not FA
                         else "بازه همین لحظه رزرو شد — دیگری انتخاب کنید.") +
                        "</div>", unsafe_allow_html=True)
        elif (not is_owner and
              len([r for r in st.session_state.reservations
                   if r["email"]==actual_email and r["status"]=="active"]) >= 5):
            st.markdown('<div class="msg-err">⚠ ' +
                        ("Max 5 active bookings per account." if not FA
                         else "حداکثر ۵ رزرو فعال برای هر حساب.") +
                        "</div>", unsafe_allow_html=True)
        else:
            sl  = slot_label(start_hr, dur)
            new = {
                "id": gen_id(), "name": actual_name, "email": actual_email, "phone": actual_phone,
                "hall": hall_key, "hall_en": hall["name_en"], "hall_fa": hall["name_fa"],
                "hall_color": hall["color"],
                "date": date_str, "start_hour": start_hr, "duration": dur,
                "slot_label": sl, "price": total, "status": "active",
                "booked_by": "owner" if is_owner else "user",
                "booked_at": datetime.now().strftime("%Y-%m-%d %H:%M"),
            }
            st.session_state.reservations.append(new)
            st.markdown(f'<div class="msg-ok">✓ ' +
                        (f"Confirmed! {hall_nm} · {sl}" if not FA
                         else f"تأیید شد! {hall_nm} · {sl}") +
                        "</div>", unsafe_allow_html=True)
            st.balloons()
            time.sleep(3)
            st.rerun()


# ─────────────────────────────────────────────
#  RES CARDS
# ─────────────────────────────────────────────
def _cell(label, value):
    """Render one metadata cell — avoids nested f-string quote conflicts."""
    return (
        '<div>'
        '<div style="font-size:9px;font-weight:600;text-transform:uppercase;color:#9b7250">'
        + label +
        '</div>'
        '<div style="margin-top:2px;font-weight:500;color:#4b3932">'
        + str(value) +
        '</div>'
        '</div>'
    )

def res_cards(res_list, pfx, show_email=False):
    FA = is_fa()
    for r in sorted(res_list, key=lambda x: x["date"], reverse=True):
        cancelled = r["status"] == "cancelled"
        hn        = r["hall_fa"] if FA else r["hall_en"]
        badge_cls = "badge-cancelled" if cancelled else "badge-active"
        badge_txt = ("✕ Cancelled" if not FA else "✕ لغو شده") if cancelled else ("✓ Confirmed" if not FA else "✓ تأیید شده")
        owner_tag = ('<span class="badge-owner">👑 ' + ("Owner" if not FA else "مدیر") + '</span>'
                     if r.get("booked_by") == "owner" else "")
        card_cls  = "res-card cancelled" if cancelled else "res-card"
        cols      = 5 if show_email else 4

        # Cancellation policy note
        if not cancelled:
            fee, is_free = cancellation_fee(r)
            if is_free:
                note = ("✓ Free cancellation — more than 48h before start." if not FA
                        else "✓ لغو رایگان — بیش از ۴۸ ساعت تا شروع رزرو.")
                note_color = "#2d6b4f"
            else:
                note = ("⚠ Late cancellation — 100% fee applies (less than 48h to start)." if not FA
                        else "⚠ لغو دیر هنگام — ۱۰۰٪ مبلغ دریافت می‌شود (کمتر از ۴۸ ساعت تا شروع).")
                note_color = "#8b2d36"
        else:
            note = ("Cancelled." if not FA else "لغو شده.")
            note_color = "#9b7250"

        meta_guest = _cell("Guest" if not FA else "مهمان", r["name"])
        meta_email = _cell("Email", r["email"]) if show_email else ""
        meta_date  = _cell("Date"  if not FA else "تاریخ",  r["date"])
        meta_slot  = _cell("Slot"  if not FA else "زمان",   r["slot_label"])
        meta_dur   = _cell("Dur."  if not FA else "بازه",    str(r["duration"]) + "h")

        html = (
            '<div class="' + card_cls + '">'
            '<div style="display:flex;justify-content:space-between;align-items:flex-start;gap:10px;flex-wrap:wrap">'
            '<div>'
            '<div class="res-id">' + r["id"] + '</div>'
            '<div class="res-hall">' + hn + '</div>'
            '<span class="' + badge_cls + '">' + badge_txt + '</span> ' + owner_tag +
            '</div>'
            '<div class="res-price">' + fmt(r["price"]) + '</div>'
            '</div>'
            '<div style="display:grid;grid-template-columns:repeat(' + str(cols) + ',1fr);gap:8px;margin-top:10px;font-size:12px">'
            + meta_guest + meta_email + meta_date + meta_slot + meta_dur +
            '</div>'
            '<div style="margin-top:10px;padding-top:8px;border-top:1px solid #eee0d3;font-size:11px;color:' + note_color + '">'
            + note +
            '</div>'
            '</div>'
        )
        st.markdown(html, unsafe_allow_html=True)

        if not cancelled:
            fee, is_free = cancellation_fee(r)
            confirm_key = f"confirm_cancel_{r['id']}"
            if st.session_state.get(confirm_key, False):
                # Show confirmation with fee warning
                if not is_free:
                    st.markdown(
                        '<div class="msg-err">⚠ ' +
                        (f"You will be charged the full amount ({fmt(fee)}) — less than 48h to start. Confirm?"
                         if not FA else
                         f"مبلغ کامل ({fmt(fee)}) دریافت می‌شود — کمتر از ۴۸ ساعت تا شروع. تأیید می‌کنید؟") +
                        '</div>',
                        unsafe_allow_html=True
                    )
                bcc1, bcc2 = st.columns(2)
                with bcc1:
                    st.markdown('<div class="btn-cancel">', unsafe_allow_html=True)
                    if st.button(
                        ("⚠ Confirm Cancel" if not FA else "⚠ تأیید لغو") + (" (full charge)" if not is_free else " (free)"),
                        key=f"{pfx}_cyes_{r['id']}"
                    ):
                        cancel_res(r["id"])
                        st.session_state[confirm_key] = False
                        st.rerun()
                    st.markdown("</div>", unsafe_allow_html=True)
                with bcc2:
                    st.markdown('<div class="btn-ghost">', unsafe_allow_html=True)
                    if st.button("← Keep" if not FA else "← نگه‌داشتن", key=f"{pfx}_cno_{r['id']}"):
                        st.session_state[confirm_key] = False
                        st.rerun()
                    st.markdown("</div>", unsafe_allow_html=True)
            else:
                st.markdown('<div class="btn-cancel">', unsafe_allow_html=True)
                if st.button(
                    ("Cancel & Refund" if is_free else "Cancel (charge applies)") if not FA
                    else ("لغو و بازپرداخت" if is_free else "لغو (هزینه دریافت می‌شود)"),
                    key=f"{pfx}_c_{r['id']}"
                ):
                    st.session_state[confirm_key] = True
                    st.rerun()
                st.markdown("</div>", unsafe_allow_html=True)


# ══════════════════════════════════════════════
#  PAGES
# ══════════════════════════════════════════════

# ──────────────────────────────
#  GATEWAY
# ──────────────────────────────
if st.session_state.page == "gateway":
    FA = is_fa()
    st.markdown(f"""
    <div class="hero" style="text-align:center;padding:52px 32px">
      <div style="font-size:52px;margin-bottom:10px">🎭</div>
      <h1>Padmart</h1>
      <p style="font-size:13px;letter-spacing:.25em;text-transform:uppercase;
                color:rgba(200,169,110,.9);margin-top:4px">
        {"Acting School · Hall Reservations" if not FA else "مدرسه بازیگری · رزرو پلاتو"}
      </p>
    </div>
    """, unsafe_allow_html=True)

    c1, c2, c3 = st.columns(3, gap="medium")
    with c1:
        st.markdown(f"""<div class="gcard">
          <div style="font-size:40px;margin-bottom:12px">🆕</div>
          <div style="font-family:'Playfair Display',serif;font-size:20px;color:var(--wine);margin-bottom:8px">
            {"Register" if not FA else "ثبت‌نام"}</div>
          <div style="font-size:13px;color:var(--muted);line-height:1.65">
            {"Create an account to book halls and track your full history"
             if not FA else "ثبت‌نام کن تا هم پلاتو رزرو کنی، هم به همه سوابق رزروهات دسترسی داشته باشی"}</div>
        </div>""", unsafe_allow_html=True)
        if st.button("Create Account →" if not FA else "ایجاد حساب ←", key="b_reg"):
            nav("register")
    with c2:
        st.markdown(f"""<div class="gcard">
          <div style="font-size:40px;margin-bottom:12px">🎟</div>
          <div style="font-family:'Playfair Display',serif;font-size:20px;color:var(--wine);margin-bottom:8px">
            {"Sign In" if not FA else "ورود"}</div>
          <div style="font-size:13px;color:var(--muted);line-height:1.65">
            {"Sign in to book, view your calendar and manage reservations"
             if not FA else "وارد حسابت شو تا رزرو کنی، جدول زمان‌بندیت رو ببینی و رزروهات رو مدیریت کنی"}</div>
        </div>""", unsafe_allow_html=True)
        if st.button("Sign In →" if not FA else "ورود ←", key="b_login"):
            nav("login")
    with c3:
        st.markdown(f"""<div class="gcard-owner">
          <div style="font-size:40px;margin-bottom:12px">👑</div>
          <div style="font-family:'Playfair Display',serif;font-size:20px;color:#fff;margin-bottom:8px">
            {"Site Owner" if not FA else "مدیر سایت"}</div>
          <div style="font-size:13px;color:rgba(255,255,255,.62);line-height:1.65">
            {"Full dashboard — all reservations, analytics & settings"
             if not FA else "داشبورد کامل — همه رزروها، آمار و تنظیمات"}</div>
        </div>""", unsafe_allow_html=True)
        st.markdown('<div class="btn-gold">', unsafe_allow_html=True)
        if st.button("Owner Login →" if not FA else "ورود مدیر ←", key="b_owner"):
            nav("owner-login")
        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown(f"""<p style="text-align:center;font-size:13px;color:var(--muted);margin-top:20px">
      {"Grand Studio · Studio 2 · 70,000 تومان/hr · Full Refund on Cancellation"
       if not FA else "پلاتو بزرگ · استودیو ۲ · ۷۰،۰۰۰ تومان/ساعت · بازپرداخت کامل در صورت لغو"}
    </p>""", unsafe_allow_html=True)


# ──────────────────────────────
#  REGISTER
# ──────────────────────────────
elif st.session_state.page == "register":
    FA = is_fa()
    st.markdown('<div class="btn-ghost">', unsafe_allow_html=True)
    if st.button("← Back" if not FA else "← بازگشت", key="reg_b"): nav("gateway")
    st.markdown("</div>", unsafe_allow_html=True)
    _, mid, _ = st.columns([1,1.4,1])
    with mid:
        st.markdown(f"""<div style="text-align:center;margin:20px 0 22px">
          <div style="font-size:40px;margin-bottom:8px">🆕</div>
          <h2 style="font-family:'Playfair Display',serif;font-size:26px;color:var(--wine)">
            {"Create Your Account" if not FA else "ایجاد حساب کاربری"}</h2>
        </div>""", unsafe_allow_html=True)
        with st.form("reg_form"):
            rn = st.text_input("Full Name"         if not FA else "نام و نام خانوادگی", placeholder="Sara Mohammadi")
            rph= st.text_input("Phone Number"       if not FA else "شماره تلفن",          placeholder="+98 912 345 6789")
            re = st.text_input("Email"             if not FA else "ایمیل",               placeholder="sara@email.com")
            rp1= st.text_input("Password"          if not FA else "رمز عبور",            type="password")
            rp2= st.text_input("Confirm Password"  if not FA else "تأیید رمز عبور",      type="password")
            sub= st.form_submit_button("Create Account" if not FA else "ایجاد حساب", use_container_width=True)
        if sub:
            ek = re.strip().lower()
            if not rn.strip():
                st.markdown('<div class="msg-err">⚠ ' + ("Enter your name." if not FA else "نام خود را وارد کنید.") + "</div>", unsafe_allow_html=True)
            elif not rph.strip():
                st.markdown('<div class="msg-err">⚠ ' + ("Enter your phone number." if not FA else "شماره تلفن خود را وارد کنید.") + "</div>", unsafe_allow_html=True)
            elif not ek or "@" not in ek:
                st.markdown('<div class="msg-err">⚠ ' + ("Enter a valid email." if not FA else "ایمیل معتبر وارد کنید.") + "</div>", unsafe_allow_html=True)
            elif ek in st.session_state.users:
                st.markdown('<div class="msg-err">⚠ ' + ("Email already registered." if not FA else "این ایمیل قبلاً ثبت‌نام شده.") + "</div>", unsafe_allow_html=True)
            elif len(rp1) < 6:
                st.markdown('<div class="msg-err">⚠ ' + ("Password min 6 characters." if not FA else "رمز عبور حداقل ۶ کاراکتر.") + "</div>", unsafe_allow_html=True)
            elif rp1 != rp2:
                st.markdown('<div class="msg-err">⚠ ' + ("Passwords do not match." if not FA else "رمزها یکسان نیستند.") + "</div>", unsafe_allow_html=True)
            else:
                st.session_state.users[ek] = {"name": rn.strip(), "phone": rph.strip(), "email": ek, "pass_hash": hpw(rp1)}
                st.session_state.current_user = ek
                nav("home")


# ──────────────────────────────
#  LOGIN
# ──────────────────────────────
elif st.session_state.page == "login":
    FA = is_fa()
    st.markdown('<div class="btn-ghost">', unsafe_allow_html=True)
    if st.button("← Back" if not FA else "← بازگشت", key="login_b"): nav("gateway")
    st.markdown("</div>", unsafe_allow_html=True)
    _, mid, _ = st.columns([1,1.4,1])
    with mid:
        st.markdown(f"""<div style="text-align:center;margin:20px 0 22px">
          <div style="font-size:40px;margin-bottom:8px">🎟</div>
          <h2 style="font-family:'Playfair Display',serif;font-size:26px;color:var(--wine)">
            {"Sign In" if not FA else "ورود به حساب"}</h2>
        </div>""", unsafe_allow_html=True)
        with st.form("login_form"):
            le  = st.text_input("Email"    if not FA else "ایمیل",    placeholder="sara@email.com")
            lp  = st.text_input("Password" if not FA else "رمز عبور", type="password")
            sub = st.form_submit_button("Sign In" if not FA else "ورود", use_container_width=True)
        if sub:
            k = le.strip().lower()
            u = st.session_state.users.get(k)
            if not u or u["pass_hash"] != hpw(lp):
                st.markdown('<div class="msg-err">⚠ ' + ("Wrong email or password." if not FA else "ایمیل یا رمز عبور نادرست.") + "</div>", unsafe_allow_html=True)
            else:
                st.session_state.current_user = k
                nav("home")
        st.markdown("---")
        st.markdown(f'<p style="text-align:center;font-size:13px;color:var(--muted)">{"No account?" if not FA else "حساب ندارید؟"}</p>', unsafe_allow_html=True)
        if st.button("Create Account" if not FA else "ثبت‌نام کنید", key="l_reg"): nav("register")


# ──────────────────────────────
#  USER HOME
# ──────────────────────────────
elif st.session_state.page == "home":
    FA  = is_fa()
    cu  = st.session_state.current_user
    if not cu or cu not in st.session_state.users:
        nav("login")

    user   = st.session_state.users[cu]
    my_res = [r for r in st.session_state.reservations if r["email"] == cu]
    all_active = [r for r in st.session_state.reservations if r["status"] == "active"]

    # nav bar
    nc1, nc2 = st.columns([5,1])
    with nc2:
        st.markdown('<div class="btn-ghost">', unsafe_allow_html=True)
        if st.button("Sign Out" if not FA else "خروج", key="h_out"):
            st.session_state.current_user = None
            nav("gateway")
        st.markdown("</div>", unsafe_allow_html=True)

    # hero
    st.markdown(f"""
    <div class="hero">
      <div class="hero-badge">🎟 {"Welcome back" if not FA else "خوش آمدید"}</div>
      <h1>{user['name']}</h1>
      <p>{"Book a hall, check the live availability calendar and manage your reservations."
          if not FA else "وضعیت ظرفیت‌های خالی رو به‌صورت آنلاین و لحظه‌ای بررسی کن و رزروهات رو مدیریت کن"}</p>
    </div>
    """, unsafe_allow_html=True)

    # quick stats
    my_active = [r for r in my_res if r["status"]=="active"]
    s1,s2,s3 = st.columns(3)
    s1.markdown(f'<div class="stat-card"><div class="stat-num">{len(my_res)}</div><div class="stat-lbl">{"Total Bookings" if not FA else "کل رزروها"}</div></div>', unsafe_allow_html=True)
    s2.markdown(f'<div class="stat-card"><div class="stat-num">{len(my_active)}</div><div class="stat-lbl">{"Active" if not FA else "فعال"}</div></div>', unsafe_allow_html=True)
    s3.markdown(f'<div class="stat-card"><div class="stat-num">{fmt(sum(r["price"] for r in my_active))}</div><div class="stat-lbl">{"Total Spent" if not FA else "مجموع هزینه"}</div></div>', unsafe_allow_html=True)

    st.markdown("")

    tab_book, tab_cal, tab_myres = st.tabs([
        "📅 " + ("Book a Hall"              if not FA else "رزرو سالن"),
        "📆 " + ("Availability Calendar"    if not FA else "تقویم موجودی"),
        "📋 " + ("My Account & Reservations" if not FA else "حساب و رزروهای من"),
    ])

    # ── Book ──
    with tab_book:
        left, right = st.columns([1.3, 0.85], gap="medium")
        with left:
            st.markdown(f"""
            <div class="eyebrow">{"Book a Rehearsal Hall" if not FA else "رزرو سالن تمرین"}</div>
            <h2 style="font-family:'Playfair Display',serif;font-size:22px;margin-bottom:16px">
              {"Choose your hall, duration and time slot" if not FA
               else "سالن، مدت و بازه زمانی را انتخاب کنید"}
            </h2>
            """, unsafe_allow_html=True)
            booking_form("user", cu, user["name"], is_owner=False)
        with right:
            st.markdown(f"""
            <div class="panel-dark">
              <div class="eyebrow eyebrow-light">{"How It Works" if not FA else "نحوه عملکرد"}</div>
              <div style="font-size:17px;font-weight:700;font-family:'Playfair Display',serif;margin:6px 0 14px">
                {"Simple Steps" if not FA else "مراحل ساده"}</div>
              <div style="font-size:12px;color:rgba(255,255,255,.72);line-height:1.7">
                <div style="display:flex;gap:8px;margin-bottom:8px"><span style="color:var(--gold);font-weight:700;flex-shrink:0">01</span>
                  <span>{"Check the Availability Calendar tab first." if not FA else "ابتدا جدول زمان بندی را بررسی کنید."}</span></div>
                <div style="display:flex;gap:8px;margin-bottom:8px"><span style="color:var(--gold);font-weight:700;flex-shrink:0">02</span>
                  <span>{"Select a hall and set your duration." if not FA else "پلاتو را انتخاب کنید و بازه زمان رزرو را انتخاب کنید."}</span></div>
                <div style="display:flex;gap:8px;margin-bottom:8px"><span style="color:var(--gold);font-weight:700;flex-shrink:0">03</span>
                  <span>{"Only conflict-free slots are offered." if not FA else "فقط بازه‌های بدون تداخل پیشنهاد می‌شود."}</span></div>
                <div style="display:flex;gap:8px"><span style="color:var(--gold);font-weight:700;flex-shrink:0">04</span>
                  <span>{"Free cancellation up to 48 hours before your reservation. After that, the full amount will be charged.
" if not FA else "لغو رزرو تا ۴۸ ساعت قبل رایگان است. پس از آن، مبلغ کامل دریافت می‌شود."}</span></div>
              </div>
            </div>
            <div class="panel">
              <div class="eyebrow" style="margin-bottom:10px">{"Hall Prices" if not FA else "قیمت پلاتوها"}</div>
              <div style="display:flex;gap:10px;margin-bottom:12px">
                <div style="font-size:24px">🎪</div>
                <div>
                  <div style="font-weight:600">{"Grand Studio" if not FA else "پلاتو بزرگ"}</div>
                  <div style="font-size:11px;color:var(--muted)">120 seats · Pro lighting</div>
                  <div style="font-size:14px;font-weight:700;color:var(--wine)">70,000 تومان/hr</div>
                </div>
              </div>
              <div style="display:flex;gap:10px">
                <div style="font-size:24px">🎭</div>
                <div>
                  <div style="font-weight:600">{"Studio 2" if not FA else "پلاتو ۲"}</div>
                  <div style="font-size:11px;color:var(--muted)">30 seats · Mirrors</div>
                  <div style="font-size:14px;font-weight:700;color:var(--wine)">70,000 تومان/hr</div>
                </div>
              </div>
            </div>
            """, unsafe_allow_html=True)

    # ── Availability Calendar (same as owner — shows ALL bookings) ──
    with tab_cal:
        FA = is_fa()
        st.markdown(
            '<div class="eyebrow">' + ("Live Availability" if not FA else "زمان‌های قابل رزرو") + '</div>'
            '<h2 style="font-family:\'Playfair Display\',serif;font-size:22px;margin-bottom:4px">'
            + ("Hall Availability Calendar" if not FA else "جدول زمان‌های قابل رزرو") +
            '</h2>'
            '<p style="font-size:13px;color:var(--muted);margin-bottom:16px">'
            + ("Select a hall — cream cells are free. Pick a date, set duration and book directly from the calendar. Your bookings appear in gold."
               if not FA else
               "یک پلاتو انتخاب کنید — وقت های خالی کرم رنگ هستند. تاریخ و بازه زمانی را انتخاب کنید و مستقیم از تقویم رزرو کنید. رزروهای شما با طلایی نشان داده می‌شود")
            + '</p>',
            unsafe_allow_html=True
        )

        sel_hall = st.radio(
            "Hall" if not FA else "پلاتو",
            options=list(HALLS.keys()),
            format_func=lambda k: HALLS[k]["icon"] + " " + (HALLS[k]["name_fa"] if FA else HALLS[k]["name_en"]),
            horizontal=True,
            key="u_cal_hall"
        )
        days = st.slider(
            "Days ahead" if not FA else "روزهای آینده",
            7, 30, 14, key="u_cal_days"
        )
        hall_calendar(
            sel_hall,
            days_ahead=days,
            highlight_email=cu,
            allow_booking=True,
            booking_email=cu,
            booking_name=user["name"],
            is_owner=False,
            cal_prefix="user",
        )

    # ── My Account & Reservations ──
    with tab_myres:
        FA = is_fa()
        st.markdown(f"""
        <div class="eyebrow">{"Account & Reservations" if not FA else "حساب و رزروها"}</div>
        <h2 style="font-family:'Playfair Display',serif;font-size:22px;margin-bottom:16px">
          {user['name']}</h2>
        """, unsafe_allow_html=True)

        st.markdown(f"""
        <div class="panel" style="margin-bottom:18px">
          <div style="display:grid;grid-template-columns:1fr 1fr;gap:14px;font-size:13px">
            <div><div style="font-size:10px;font-weight:600;text-transform:uppercase;color:#9b7250">{"Name" if not FA else "نام"}</div>
              <div style="font-weight:600;margin-top:3px">{user['name']}</div></div>
            <div><div style="font-size:10px;font-weight:600;text-transform:uppercase;color:#9b7250">{"Phone" if not FA else "تلفن"}</div>
              <div style="font-weight:600;margin-top:3px">{user.get('phone','—')}</div></div>
            <div><div style="font-size:10px;font-weight:600;text-transform:uppercase;color:#9b7250">Email</div>
              <div style="font-weight:600;margin-top:3px">{cu}</div></div>
            <div><div style="font-size:10px;font-weight:600;text-transform:uppercase;color:#9b7250">{"Total Bookings" if not FA else "کل رزروها"}</div>
              <div style="font-weight:600;color:var(--wine);margin-top:3px">{len(my_res)}</div></div>
            <div><div style="font-size:10px;font-weight:600;text-transform:uppercase;color:#9b7250">{"Total Spent" if not FA else "مجموع هزینه"}</div>
              <div style="font-weight:600;color:var(--wine);margin-top:3px">{fmt(sum(r['price'] for r in my_active))}</div></div>
          </div>
        </div>
        """, unsafe_allow_html=True)

        if not my_res:
            st.markdown(f"""<div style="text-align:center;padding:40px;background:var(--paper);
              border:2px dashed #d7c5b4;border-radius:20px">
              <div style="font-size:34px;margin-bottom:10px">🎟</div>
              <div style="font-family:'Playfair Display',serif;font-size:18px">
                {"No reservations yet" if not FA else "هنوز رزروی ندارید"}</div></div>""",
                unsafe_allow_html=True)
        else:
            res_cards(my_res, "user_res", show_email=False)

        with st.expander("📞 " + ("Update Phone Number" if not FA else "تغییر شماره تلفن")):
            with st.form("u_phone_form"):
                new_phone = st.text_input(
                    "New Phone Number" if not FA else "شماره تلفن جدید",
                    value=user.get("phone",""),
                    placeholder="+98 912 345 6789"
                )
                if st.form_submit_button("Save" if not FA else "ذخیره", use_container_width=True):
                    if not new_phone.strip():
                        st.error("Enter a phone number." if not FA else "شماره تلفن را وارد کنید.")
                    else:
                        st.session_state.users[cu]["phone"] = new_phone.strip()
                        st.success("Phone updated!" if not FA else "شماره تلفن به‌روز شد!")

        with st.expander("🔒 " + ("Change Password" if not FA else "تغییر رمز عبور")):
            with st.form("u_pw_form"):
                op  = st.text_input("Current Password"     if not FA else "رمز فعلی",    type="password")
                np1 = st.text_input("New Password"          if not FA else "رمز جدید",    type="password")
                np2 = st.text_input("Confirm New Password"  if not FA else "تأیید رمز جدید", type="password")
                if st.form_submit_button("Save" if not FA else "ذخیره", use_container_width=True):
                    if user["pass_hash"] != hpw(op):
                        st.error("Current password incorrect." if not FA else "رمز فعلی نادرست.")
                    elif len(np1) < 6:
                        st.error("Min 6 characters." if not FA else "حداقل ۶ کاراکتر.")
                    elif np1 != np2:
                        st.error("Passwords do not match." if not FA else "رمزها یکسان نیستند.")
                    else:
                        st.session_state.users[cu]["pass_hash"] = hpw(np1)
                        st.success("Password updated!" if not FA else "رمز عبور به‌روز شد!")


# ──────────────────────────────
#  OWNER LOGIN
# ──────────────────────────────
elif st.session_state.page == "owner-login":
    FA = is_fa()
    st.markdown('<div class="btn-ghost">', unsafe_allow_html=True)
    if st.button("← Back" if not FA else "← بازگشت", key="ol_b"): nav("gateway")
    st.markdown("</div>", unsafe_allow_html=True)
    _, mid, _ = st.columns([1,1.2,1])
    with mid:
        st.markdown(f"""<div style="text-align:center;margin:28px 0 22px">
          <div style="font-size:40px;margin-bottom:8px">👑</div>
          <h2 style="font-family:'Playfair Display',serif;font-size:26px;color:var(--wine)">
            {"Owner Login" if not FA else "ورود مدیر"}</h2>
        </div>""", unsafe_allow_html=True)
        with st.form("ol_form"):
            pw  = st.text_input("Password" if not FA else "رمز عبور", type="password")
            sub = st.form_submit_button("Enter Dashboard" if not FA else "ورود به داشبورد", use_container_width=True)
        if sub:
            if pw == st.session_state.owner_pass:
                st.session_state.owner_auth = True
                nav("owner")
            else:
                st.markdown('<div class="msg-err">⚠ ' + ("Incorrect password." if not FA else "رمز عبور نادرست.") + "</div>", unsafe_allow_html=True)
        st.markdown(f'<p style="text-align:center;font-size:12px;color:var(--muted);margin-top:12px">{"Demo:" if not FA else " نمونه:"} <strong>admin123</strong></p>', unsafe_allow_html=True)


# ──────────────────────────────
#  OWNER DASHBOARD
# ──────────────────────────────
elif st.session_state.page == "owner":
    FA = is_fa()
    if not st.session_state.owner_auth:
        nav("owner-login")

    oc1, oc2 = st.columns([5,1])
    with oc2:
        st.markdown('<div class="btn-ghost">', unsafe_allow_html=True)
        if st.button("← Exit" if not FA else "← خروج", key="o_out"):
            st.session_state.owner_auth = False
            nav("gateway")
        st.markdown("</div>", unsafe_allow_html=True)

    st.markdown(f"""
    <div class="owner-hero">
      <div class="owner-badge">👑 {"Owner Dashboard" if not FA else "داشبورد مدیر"}</div>
      <h1>{"Full Control. All Reservations." if not FA else "کنترل کامل. همه رزروها."}</h1>
      <p>{"Book for guests, manage all reservations, view the full calendar and analytics."
          if not FA else "برای مهمانان رزرو کنید، همه رزروها را مدیریت کنید، تقویم کامل و آمار را ببینید."}</p>
    </div>
    """, unsafe_allow_html=True)

    all_res    = st.session_state.reservations
    active_res = [r for r in all_res if r["status"]=="active"]
    revenue    = sum(r["price"] for r in active_res)
    ug         = len(set(r["email"] for r in all_res))

    s1,s2,s3,s4 = st.columns(4)
    for col, val, lbl in [
        (s1, len(all_res),   "Total Bookings" if not FA else "کل رزروها"),
        (s2, len(active_res),"Active"          if not FA else "فعال"),
        (s3, f"{revenue:,}", "Revenue (تومان)" if not FA else "درآمد (تومان)"),
        (s4, ug,             "Unique Guests"   if not FA else "مهمانان غیرعضو"),
    ]:
        col.markdown(f'<div class="stat-card"><div class="stat-num">{val}</div><div class="stat-lbl">{lbl}</div></div>', unsafe_allow_html=True)

    st.markdown("")

    tb1, tb2, tb3, tb4, tb5, tb6 = st.tabs([
        "📅 " + ("Book for Guest"    if not FA else "رزرو برای مهمان"),
        "📆 " + ("Full Calendar"     if not FA else "تقویم کامل"),
        "📋 " + ("All Reservations"  if not FA else "همه رزروها"),
        "📊 " + ("Analytics"         if not FA else "آمار"),
        "👥 " + ("Manage Users"      if not FA else "مدیریت کاربران"),
        "⚙️ " + ("Settings"          if not FA else "تنظیمات"),
    ])

    # ── Owner: Book for Guest ──
    with tb1:
        lo, ro = st.columns([1.3, 0.85], gap="medium")
        with lo:
            st.markdown(f"""<div class="eyebrow">{"Owner Booking" if not FA else "رزرو مدیر"}</div>
            <h2 style="font-family:'Playfair Display',serif;font-size:22px;margin-bottom:16px">
              {"Reserve a Hall on Behalf of a Guest" if not FA else "رزرو پلاتو به نمایندگی از مهمان"}</h2>""",
                        unsafe_allow_html=True)
            booking_form("owner", "owner@padmart.internal", "Owner", is_owner=True)
        with ro:
            st.markdown(f"""<div class="panel-dark">
              <div class="eyebrow eyebrow-light">{"Owner Booking Note" if not FA else "یادداشت مربوط به رزرو"}</div>
              <div style="font-size:13px;color:rgba(255,255,255,.75);line-height:1.7;margin-top:8px">
                {"As owner you can book any hall for any guest. The reservation will appear on the "
                 "calendar and in All Reservations."
                 if not FA else
                 "به عنوان مدیر می‌توانید هر پلاتویی را برای هر مهمانی رزرو کنید. رزرو در تقویم و لیست همه رزروها نمایش داده می‌شود."}
              </div></div>""", unsafe_allow_html=True)

    # ── Owner: Full Calendar ──
    with tb2:
        FA = is_fa()
        st.markdown(
            '<div class="eyebrow">' + ("Full Reservation Calendar" if not FA else "تقویم کامل رزروها") + '</div>'
            '<h2 style="font-family:\'Playfair Display\',serif;font-size:22px;margin-bottom:4px">'
            + ("All Studios · Book Directly from Calendar" if not FA else "همه پلاتوها · رزرو مستقیم از تقویم") +
            '</h2>'
            '<p style="font-size:13px;color:var(--muted);margin-bottom:16px">'
            + ("Select a hall — X axis shows days (above), Y axis shows hours. Cream = free. You can book directly below the chart."
               if not FA else
               "یک سالن انتخاب کنید — محور X روزها (بالا) و محور Y ساعات است. کرم = آزاد. می‌توانید مستقیماً زیر نمودار رزرو کنید.")
            + '</p>',
            unsafe_allow_html=True
        )
        o_hall = st.radio(
            "Hall" if not FA else "پلاتو",
            options=list(HALLS.keys()),
            format_func=lambda k: HALLS[k]["icon"] + " " + (HALLS[k]["name_fa"] if FA else HALLS[k]["name_en"]),
            horizontal=True, key="o_cal_hall"
        )
        o_days = st.slider("Days ahead" if not FA else "روزهای آینده", 7, 30, 14, key="o_cal_days")
        hall_calendar(
            o_hall,
            days_ahead=o_days,
            highlight_email=None,
            allow_booking=True,
            booking_email="owner@padmart.internal",
            booking_name="Owner",
            is_owner=True,
            cal_prefix="owner",
        )

    # ── Owner: All Reservations ──
    with tb3:
        FA = is_fa()
        # Excel export
        ex_col1, ex_col2 = st.columns([3,1])
        with ex_col1:
            srch = st.text_input("🔍", placeholder="Search by name or email…" if not FA else "جستجو با نام یا ایمیل…",
                                 label_visibility="collapsed", key="o_srch")
        with ex_col2:
            if EXCEL_OK and all_res:
                xls_buf = build_excel(all_res, st.session_state.users)
                st.download_button(
                    label="📥 " + ("Export Excel" if not FA else "خروجی اکسل"),
                    data=xls_buf,
                    file_name="padmart_reservations.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            elif not EXCEL_OK:
                st.caption("pip install openpyxl for Excel export")
        filt = [r for r in sorted(all_res, key=lambda x: x["date"], reverse=True)
                if not srch or srch.lower() in r["name"].lower() or srch.lower() in r["email"]]
        if not filt:
            st.markdown(f"""<div style="text-align:center;padding:44px;background:var(--paper);
              border:2px dashed #d7c5b4;border-radius:20px">
              <div style="font-size:34px;margin-bottom:10px">🎟</div>
              <div style="font-family:'Playfair Display',serif;font-size:18px">
                {"No reservations yet" if not FA else "هنوز رزروی ثبت نشده"}</div></div>""",
                unsafe_allow_html=True)
        else:
            res_cards(filt, "o_all", show_email=True)

    # ── Owner: Analytics ──
    with tb4:
        FA = is_fa()
        from collections import Counter
        ac1, ac2, ac3 = st.columns(3)

        CHART_TICK = "#241817"   # dark ink — clearly readable on cream
        CHART_GRID = "rgba(200,180,160,.45)"

        with ac1:
            nc = Counter(r["name"] for r in active_res)
            fig = pgo.Figure(pgo.Bar(
                x=list(nc.keys()), y=list(nc.values()),
                marker=dict(
                    color=list(nc.values()),
                    colorscale=[[0, CREAM], [0.5, GOLD], [1, WINE]],
                    line=dict(color=WINE, width=1)
                ),
                text=list(nc.values()),
                textposition="outside",
                textfont=dict(color=CHART_TICK, size=12, family="DM Sans"),
                hovertemplate="<b>%{x}</b><br>Bookings: %{y}<extra></extra>",
            ))
            fig.update_layout(
                title=dict(
                    text="Bookings by Guest" if not FA else "رزرو به تفکیک اعضا",
                    font=dict(size=14, color=WINE, family="Playfair Display, serif"), x=.5
                ),
                plot_bgcolor=CREAM, paper_bgcolor=CREAM,
                xaxis=dict(
                    showgrid=False,
                    tickfont=dict(size=12, color=CHART_TICK, family="DM Sans"),
                    title=dict(text=""),
                ),
                yaxis=dict(
                    gridcolor=CHART_GRID,
                    tickfont=dict(size=12, color=CHART_TICK, family="DM Sans"),
                    title=dict(text=""),
                ),
                font=dict(color=CHART_TICK, family="DM Sans"),
                margin=dict(t=50, b=20, l=10, r=10), height=280,
                showlegend=False,
            )
            st.plotly_chart(fig, use_container_width=True)

        with ac2:
            gc = sum(1 for r in active_res if r["hall"] == "grand")
            sc = sum(1 for r in active_res if r["hall"] == "studio2")
            fig = pgo.Figure(pgo.Pie(
                labels=[
                    "Grand Studio" if not FA else "پلاتو بزرگ",
                    "Studio 2"     if not FA else "پلاتو ۲",
                ],
                values=[gc, sc],
                marker=dict(colors=[WINE, GOLD], line=dict(color="#fffaf3", width=3)),
                hole=0.42,
                textfont=dict(size=13, color="#ffffff", family="DM Sans"),
                textinfo="percent+label",
                hovertemplate="<b>%{label}</b><br>%{value} bookings<extra></extra>",
            ))
            fig.update_layout(
                title=dict(
                    text="Hall Usage" if not FA else "آنالیز رزرو پلاتوها",
                    font=dict(size=14, color=WINE, family="Playfair Display, serif"), x=.5
                ),
                plot_bgcolor=CREAM, paper_bgcolor=CREAM,
                legend=dict(
                    font=dict(size=12, color=CHART_TICK, family="DM Sans"),
                    bgcolor="rgba(0,0,0,0)",
                ),
                font=dict(color=CHART_TICK, family="DM Sans"),
                margin=dict(t=50, b=10, l=10, r=10), height=280,
            )
            st.plotly_chart(fig, use_container_width=True)

        with ac3:
            from collections import Counter as _C
            dc    = _C(r["date"] for r in active_res)
            dates = sorted(dc.keys())
            fig = pgo.Figure(pgo.Scatter(
                x=dates, y=[dc[d] for d in dates],
                mode="lines+markers",
                line=dict(color=WINE, width=2.5),
                marker=dict(color=WINE, size=8, line=dict(color="#fffaf3", width=2)),
                fill="tozeroy", fillcolor="rgba(109,31,42,.09)",
                hovertemplate="<b>%{x}</b><br>Bookings: %{y}<extra></extra>",
            ))
            fig.update_layout(
                title=dict(
                    text="Bookings Over Time" if not FA else "آمار رزروها به تفکیک زمان",
                    font=dict(size=14, color=WINE, family="Playfair Display, serif"), x=.5
                ),
                plot_bgcolor=CREAM, paper_bgcolor=CREAM,
                xaxis=dict(
                    showgrid=False,
                    tickfont=dict(size=11, color=CHART_TICK, family="DM Sans"),
                    title=dict(text=""),
                ),
                yaxis=dict(
                    gridcolor=CHART_GRID,
                    tickfont=dict(size=12, color=CHART_TICK, family="DM Sans"),
                    title=dict(text=""),
                ),
                font=dict(color=CHART_TICK, family="DM Sans"),
                margin=dict(t=50, b=20, l=10, r=10), height=280,
            )
            st.plotly_chart(fig, use_container_width=True)

    # ── Owner: Manage Users ──
    with tb5:
        FA = is_fa()
        users = st.session_state.users
        all_res_local = st.session_state.reservations

        st.markdown(
            '<div class="eyebrow">' + ("Registered Accounts" if not FA else "حساب‌های ثبت‌شده") + '</div>'
            '<h2 style="font-family:\'Playfair Display\',serif;font-size:22px;margin-bottom:4px">'
            + ("Manage Users" if not FA else "مدیریت کاربران") +
            '</h2>'
            '<p style="font-size:13px;color:var(--muted);margin-bottom:18px">'
            + ("View every registered user, their reservation history, and remove accounts if needed."
               if not FA else
               "تمام کاربران ثبت‌شده، تاریخچه رزرو آن‌ها را مشاهده کنید و در صورت نیاز حساب‌ها را حذف کنید.")
            + '</p>',
            unsafe_allow_html=True
        )

        # quick stats
        uc1, uc2, uc3 = st.columns(3)
        total_users = len(users)
        total_active_users = len({
            r["email"] for r in all_res_local
            if r["status"] == "active" and r["email"] in users
        })
        total_user_revenue = sum(
            r["price"] for r in all_res_local
            if r["status"] == "active" and r["email"] in users
        )
        uc1.markdown(f'<div class="stat-card"><div class="stat-num">{total_users}</div><div class="stat-lbl">{"Registered Users" if not FA else "کاربران ثبت‌شده"}</div></div>', unsafe_allow_html=True)
        uc2.markdown(f'<div class="stat-card"><div class="stat-num">{total_active_users}</div><div class="stat-lbl">{"Users w/ Active Bookings" if not FA else "کاربران با رزرو فعال"}</div></div>', unsafe_allow_html=True)
        uc3.markdown(f'<div class="stat-card"><div class="stat-num">{fmt(total_user_revenue)}</div><div class="stat-lbl">{"Revenue from Users" if not FA else "درآمد از کاربران"}</div></div>', unsafe_allow_html=True)

        st.markdown("")

        # search
        u_search = st.text_input(
            "🔍", placeholder="Search by name or email…" if not FA else "جستجو با نام یا ایمیل…",
            label_visibility="collapsed", key="u_mgmt_search"
        )

        if not users:
            st.markdown(
                '<div style="text-align:center;padding:44px;background:var(--paper);'
                'border:2px dashed #d7c5b4;border-radius:20px">'
                '<div style="font-size:34px;margin-bottom:10px">👤</div>'
                '<div style="font-family:\'Playfair Display\',serif;font-size:18px">'
                + ("No registered users yet" if not FA else "هنوز کاربری ثبت‌نام نکرده") +
                '</div></div>',
                unsafe_allow_html=True
            )
        else:
            filtered_emails = [
                ek for ek, u in users.items()
                if not u_search
                or u_search.lower() in u["name"].lower()
                or u_search.lower() in ek
            ]

            for ek in filtered_emails:
                u = users[ek]
                u_res = [r for r in all_res_local if r["email"] == ek]
                u_active = [r for r in u_res if r["status"] == "active"]
                u_cancelled = [r for r in u_res if r["status"] == "cancelled"]
                u_spent = sum(r["price"] for r in u_active)
                is_currently_signed_in = (st.session_state.current_user == ek)

                online_tag = (
                    '<span class="badge-active">' + ("● Online" if not FA else "● آنلاین") + '</span>'
                    if is_currently_signed_in else
                    '<span class="badge-cancelled" style="background:#f0e8db;color:#9b7250">'
                    + ("Offline" if not FA else "آفلاین") + '</span>'
                )

                st.markdown(
                    '<div class="res-card">'
                    '<div style="display:flex;justify-content:space-between;align-items:flex-start;gap:10px;flex-wrap:wrap">'
                    '<div>'
                    '<div class="res-hall">👤 ' + u["name"] + '</div>'
                    '<div style="font-size:12px;color:var(--muted);margin-top:2px">' + ek + '</div>'
                    '<div style="font-size:12px;color:var(--muted);margin-top:1px">📞 ' + u.get("phone","—") + '</div>'
                    '<div style="margin-top:6px">' + online_tag + '</div>'
                    '</div>'
                    '<div style="text-align:right">'
                    '<div class="res-price">' + fmt(u_spent) + '</div>'
                    '<div style="font-size:10px;color:var(--muted);margin-top:2px">'
                    + ("total spent" if not FA else "مجموع هزینه") + '</div>'
                    '</div>'
                    '</div>'
                    '<div style="display:grid;grid-template-columns:repeat(3,1fr);gap:8px;margin-top:12px;font-size:12px">'
                    + _cell("Active Bookings" if not FA else "رزرو فعال", len(u_active))
                    + _cell("Cancelled"        if not FA else "لغوشده",   len(u_cancelled))
                    + _cell("Total Bookings"   if not FA else "کل رزروها", len(u_res))
                    + '</div>'
                    '</div>',
                    unsafe_allow_html=True
                )

                # expandable details + actions
                with st.expander(
                    ("View Reservations & Manage" if not FA else "مشاهده رزروها و مدیریت")
                    + f" — {u['name']}"
                ):
                    if u_res:
                        res_cards(u_res, f"umgmt_{ek}", show_email=False)
                    else:
                        st.markdown(
                            '<p style="font-size:13px;color:var(--muted)">'
                            + ("This user has no reservations yet." if not FA else "این کاربر هنوز رزروی ندارد.")
                            + '</p>',
                            unsafe_allow_html=True
                        )

                    st.markdown("---")
                    dcol1, dcol2 = st.columns([3, 1])
                    with dcol1:
                        st.markdown(
                            '<p style="font-size:12px;color:#8b2d36">⚠ '
                            + ("Deleting this account will permanently remove the user and cancel all their active reservations. This cannot be undone."
                               if not FA else
                               "حذف این حساب به طور دائمی کاربر را حذف کرده و همه رزروهای فعال او را لغو می‌کند. این عمل قابل بازگشت نیست.")
                            + '</p>',
                            unsafe_allow_html=True
                        )
                    with dcol2:
                        st.markdown('<div class="btn-cancel">', unsafe_allow_html=True)
                        confirm_key = f"confirm_del_{ek}"
                        if st.session_state.get(confirm_key, False):
                            if st.button(
                                "⚠ " + ("Confirm Delete" if not FA else "تأیید حذف"),
                                key=f"del_yes_{ek}"
                            ):
                                # cancel all active reservations for this user
                                for r in st.session_state.reservations:
                                    if r["email"] == ek and r["status"] == "active":
                                        r["status"] = "cancelled"
                                # remove the user
                                del st.session_state.users[ek]
                                if st.session_state.current_user == ek:
                                    st.session_state.current_user = None
                                st.session_state[confirm_key] = False
                                st.rerun()
                        else:
                            if st.button(
                                "🗑 " + ("Delete User" if not FA else "حذف کاربر"),
                                key=f"del_{ek}"
                            ):
                                st.session_state[confirm_key] = True
                                st.rerun()
                        st.markdown("</div>", unsafe_allow_html=True)

    # ── Owner: Settings ──
    with tb6:
        FA = is_fa()
        set1, set2 = st.columns(2, gap="large")
        with set1:
            st.markdown(f'<div class="eyebrow">{"Security" if not FA else "امنیت"}</div>', unsafe_allow_html=True)
            st.markdown(f'<h3 style="font-family:\'Playfair Display\',serif;font-size:19px;margin-bottom:14px">{"Change Owner Password" if not FA else "تغییر رمز عبور مدیر"}</h3>', unsafe_allow_html=True)
            with st.form("o_pw_form"):
                np1 = st.text_input("New Password"      if not FA else "رمز جدید",       type="password")
                np2 = st.text_input("Confirm Password"  if not FA else "تأیید رمز جدید", type="password")
                if st.form_submit_button("Save Password" if not FA else "ذخیره رمز", use_container_width=True):
                    if len(np1)<6:   st.error("Min 6 characters."    if not FA else "حداقل ۶ کاراکتر.")
                    elif np1!=np2:   st.error("Passwords do not match." if not FA else "رمزها یکسان نیستند.")
                    else:
                        st.session_state.owner_pass = np1
                        st.success("Password updated!" if not FA else "رمز عبور به‌روز شد!")
        with set2:
            st.markdown(f'<div class="eyebrow">{"Email Settings" if not FA else "تنظیمات ایمیل"}</div>', unsafe_allow_html=True)
            st.markdown(f'<h3 style="font-family:\'Playfair Display\',serif;font-size:19px;margin-bottom:14px">{"SMTP Configuration" if not FA else "تنظیم SMTP"}</h3>', unsafe_allow_html=True)
            smtp = st.session_state.smtp
            with st.form("smtp_form"):
                sh = st.text_input("SMTP Host" if not FA else "هاست", value=smtp["host"], placeholder="smtp.gmail.com")
                sc1,sc2 = st.columns(2)
                with sc1: sp=st.number_input("Port",value=smtp["port"],min_value=1,max_value=9999)
                with sc2: su=st.text_input("Sender Email" if not FA else "ایمیل",value=smtp["user"])
                spw = st.text_input("App Password" if not FA else "رمز اپ",value=smtp["pass"],type="password")
                st.caption("Gmail: 2-Step Verification → App Passwords → Mail → Generate.")
                if st.form_submit_button("Save" if not FA else "ذخیره", use_container_width=True):
                    st.session_state.smtp={"host":sh,"port":int(sp),"user":su,"pass":spw}
                    st.success("Saved!" if not FA else "ذخیره شد!")


# ─────────────────────────────────────────────
#  FOOTER
# ─────────────────────────────────────────────
FA = is_fa()
st.markdown(f"""
<div style="border-top:1px solid var(--line);padding:18px 0;text-align:center;margin-top:36px">
  <span style="font-family:'Playfair Display',serif;font-size:15px;color:var(--wine)">Padmart</span>
  <span style="color:#c8aa92;margin:0 8px">·</span>
  <span style="font-size:13px;color:var(--muted)">
    {"Acting School · Grand Studio & Studio 2 · 70,000 تومان/hr"
     if not FA else "مدرسه بازیگری · پلاتو بزرگ و پلاتو ۲ · ۷۰،۰۰۰ تومان/ساعت"}
  </span>
</div>
""", unsafe_allow_html=True)
